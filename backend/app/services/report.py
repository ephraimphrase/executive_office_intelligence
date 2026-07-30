"""Executive reporting service."""
from datetime import date, datetime

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select


async def get_executive_summary(date_from: date, date_to: date, db: AsyncSession) -> dict:
    from app.models.decision import Decision
    from app.models.email_record import EmailRecord
    from app.models.event import Event
    from app.models.task import Task, TaskStatus

    start = datetime.combine(date_from, datetime.min.time())
    end   = datetime.combine(date_to, datetime.max.time())

    tasks_res   = await db.execute(select(Task))
    events_res  = await db.execute(select(Event).where(Event.start_datetime >= start, Event.start_datetime <= end))
    decisions_res = await db.execute(select(Decision))
    emails_res  = await db.execute(select(EmailRecord).where(EmailRecord.received_at >= start, EmailRecord.received_at <= end))

    tasks     = tasks_res.scalars().all()
    events    = events_res.scalars().all()
    decisions = decisions_res.scalars().all()
    emails    = emails_res.scalars().all()

    return {
        'period': {'from': str(date_from), 'to': str(date_to)},
        'meetings': {'total': len(events)},
        'tasks': {
            'total': len(tasks),
            'completed': sum(1 for t in tasks if t.status == TaskStatus.DONE),
            'overdue': sum(1 for t in tasks if t.status == TaskStatus.OVERDUE),
        },
        'decisions': {'total': len(decisions)},
        'emails': {'total': len(emails)},
    }

async def get_task_completion_report(db: AsyncSession) -> dict:
    from app.models.task import Task, TaskStatus
    res = await db.execute(select(Task))
    tasks = res.scalars().all()
    by_dept: dict = {}
    for t in tasks:
        dept = t.department or 'General'
        if dept not in by_dept:
            by_dept[dept] = {'total': 0, 'completed': 0, 'overdue': 0}
        by_dept[dept]['total'] += 1
        if t.status == TaskStatus.DONE:    by_dept[dept]['completed'] += 1
        if t.status == TaskStatus.OVERDUE: by_dept[dept]['overdue']   += 1
    return by_dept

async def get_meeting_statistics(date_from: date, date_to: date, db: AsyncSession) -> dict:
    from app.models.event import Event
    start = datetime.combine(date_from, datetime.min.time())
    end   = datetime.combine(date_to, datetime.max.time())
    res = await db.execute(select(Event).where(Event.start_datetime >= start, Event.start_datetime <= end))
    events = res.scalars().all()
    return {
        'total_meetings': len(events),
        'period': {'from': str(date_from), 'to': str(date_to)},
    }

async def get_commitment_tracking_report(db: AsyncSession) -> dict:
    from app.models.commitment import Commitment, CommitmentStatus
    res = await db.execute(select(Commitment))
    commitments = res.scalars().all()
    return {
        'total': len(commitments),
        'fulfilled': sum(1 for c in commitments if c.status == CommitmentStatus.FULFILLED),
        'overdue':   sum(1 for c in commitments if c.status == CommitmentStatus.OVERDUE),
        'pending':   sum(1 for c in commitments if c.status == CommitmentStatus.PENDING),
    }

async def get_email_analytics(date_from: date, date_to: date, db: AsyncSession) -> dict:
    from app.models.email_record import EmailPriority, EmailRecord
    start = datetime.combine(date_from, datetime.min.time())
    end   = datetime.combine(date_to, datetime.max.time())
    res = await db.execute(select(EmailRecord).where(EmailRecord.received_at >= start, EmailRecord.received_at <= end))
    emails = res.scalars().all()
    return {
        'total': len(emails),
        'critical': sum(1 for e in emails if e.priority == EmailPriority.CRITICAL),
        'high':     sum(1 for e in emails if e.priority == EmailPriority.HIGH),
    }

async def generate_action_register_report(db: AsyncSession) -> bytes:
    from app.models.task import Task
    from app.models.user import User
    from app.services.word_generator import WordGeneratorService

    res = await db.execute(select(Task).where(Task.status != "DONE").order_by(Task.due_date))
    task_rows = res.scalars().all()

    user_ids = {t.assigned_to for t in task_rows if t.assigned_to} | {t.owner_id for t in task_rows if t.owner_id}
    names: dict = {}
    if user_ids:
        users_res = await db.execute(select(User).where(User.id.in_(user_ids)))
        names = {u.id: u.full_name for u in users_res.scalars().all()}

    tasks = [
        {
            "title": t.title,
            "owner": names.get(t.assigned_to) or names.get(t.owner_id) or "",
            "due_date": t.due_date,
            "priority": t.priority.value if hasattr(t.priority, "value") else t.priority,
            "status": t.status.value if hasattr(t.status, "value") else t.status,
        }
        for t in task_rows
    ]

    svc = WordGeneratorService()
    return await svc.generate_action_register(tasks, db)

async def generate_decision_register_report(db: AsyncSession) -> bytes:
    from app.models.decision import Decision
    from app.services.word_generator import WordGeneratorService

    res = await db.execute(select(Decision).order_by(Decision.decision_date.desc()))
    decision_rows = res.scalars().all()

    decisions = [
        {
            "description": d.description,
            "made_by": d.made_by or "",
            "decision_date": d.decision_date,
            "status": d.status.value if hasattr(d.status, "value") else d.status,
            "responsible_person": d.responsible_person or "",
        }
        for d in decision_rows
    ]

    svc = WordGeneratorService()
    return await svc.generate_decision_register(decisions, db)

async def generate_word_schedule_report(date_from: date, date_to: date, db: AsyncSession) -> bytes:
    from app.models.event import Event
    from app.services.word_generator import WordGeneratorService

    start = datetime.combine(date_from, datetime.min.time())
    end = datetime.combine(date_to, datetime.max.time())
    result = await db.execute(
        select(Event).where(Event.start_datetime >= start, Event.start_datetime <= end)
        .order_by(Event.start_datetime)
    )
    events = [
        {
            "date": e.start_datetime.date().isoformat(),
            "start_time": e.start_datetime,
            "end_time": e.end_datetime,
            "agenda": e.title,
            "venue": e.location,
            "notes": e.notes,
        }
        for e in result.scalars().all()
    ]

    svc = WordGeneratorService()
    return await svc.generate_weekly_schedule(str(date_from), db, events=events)

async def export_report_to_file(report_data: dict, fmt: str = 'pdf') -> bytes:
    """Export any report dict to PDF or Excel."""
    if fmt == 'excel':
        import io

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'
        for row_idx, (key, value) in enumerate(report_data.items(), 1):
            ws.cell(row=row_idx, column=1, value=str(key))
            ws.cell(row=row_idx, column=2, value=str(value))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    else:
        # Simple PDF with ReportLab
        import io

        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        story = [Paragraph('Executive Report', styles['Title'])]
        for key, value in report_data.items():
            story.append(Paragraph(f'<b>{key}:</b> {value}', styles['Normal']))
        doc.build(story)
        return buf.getvalue()
